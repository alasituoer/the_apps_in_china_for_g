#coding:utf-8
import pandas as pd
from openpyxl import Workbook

# 此程序完成 将一款游戏的内购物品(含价格)记录的第一条、第二条和剩余条分别存入三个文件中

working_space = '/Users/Alas/Documents/TD/Top_Apps_in_China_For_Google/in_purchase_apps_in_itunes/'
path_in_purchase_info = working_space + 'result_in_purchase_info.txt'
df_in_purchase_info = pd.read_csv(path_in_purchase_info, header=None)
df_in_purchase_info.columns = ['appid', 'apptype', 'appname', 'title', 'price',]
print df_in_purchase_info.head()


wb = Workbook()
ws = wb.active
ws.title = 'p1'
ws = wb.create_sheet()
ws.title = 'p2'
ws = wb.create_sheet()
ws.title = 'p3'

for appid in df_in_purchase_info['appid'].unique():
    df_app = df_in_purchase_info[df_in_purchase_info['appid']==appid]
    
#    print df_app[:1]
    ws = wb['p1']
    ws.append(list(df_app[:1].values[0]))
    
    try:
#	print df_app[1:2]
	ws = wb['p2']
	ws.append(list(df_app[1:2].values[0]))
    except Exception, e:
	print e, '\t', appid, 'p2 is error'

    try:
#	print df_app[2:]
	ws = wb['p3']
	for i in range(len(df_app[2:])):
	    ws.append(list(df_app[2:][i:i+1].values[0]))
    except Exception, e:
	print e, '\t', appid, 'p3 is error'
wb.save(working_space + 'in_purchase_info_classified.xlsx')


